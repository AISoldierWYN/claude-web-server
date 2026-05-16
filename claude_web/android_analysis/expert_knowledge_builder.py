"""Project-local Android expert knowledge pack generation utilities."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Dict, List

from .expert_knowledge import DEFAULT_PROJECT_KNOWLEDGE_RELATIVE_PATH, load_project_knowledge_dir


EVIDENCE_TEMPLATE_FIELDS = [
    'id',
    'module_id',
    'subcategory_id',
    'profile',
    'log_type',
    'regex',
    'parameters',
    'code_location',
    'meaning',
    'severity',
    'time_anchor',
    'next_steps',
    'enabled',
]

XML_STATE_TEMPLATE_FIELDS = [
    'id',
    'module_id',
    'subcategory_id',
    'profile',
    'source_type',
    'path_patterns',
    'key_regex',
    'value_regex',
    'value_source',
    'code_location',
    'meaning',
    'severity',
    'time_anchor',
    'next_steps',
    'enabled',
]


def create_project_knowledge_scaffold(
    project_root: Path,
    *,
    module: Dict[str, Any],
    subcategories: List[Dict[str, Any]] | None = None,
    evidence_templates: List[Dict[str, Any]] | None = None,
    xml_state_templates: List[Dict[str, Any]] | None = None,
    relative_path: str = DEFAULT_PROJECT_KNOWLEDGE_RELATIVE_PATH,
    overwrite: bool = False,
    include_skill: bool = True,
) -> Dict[str, Any]:
    """Create a reviewable project-local Android expert knowledge pack draft."""

    project_root = Path(project_root).resolve()
    if not project_root.is_dir():
        raise ValueError(f'Project root does not exist: {project_root}')
    rel = _normalize_relative_path(relative_path)
    knowledge_dir = (project_root / rel).resolve()
    module_data = _normalize_module(module or {}, project_root, include_skill)
    module_id = module_data['id']
    subcategory_items = _normalize_subcategories(subcategories or [], module_id)
    evidence_items = [_normalize_evidence_template(item, module_id) for item in (evidence_templates or [])]
    xml_state_items = [_normalize_xml_state_template(item, module_id) for item in (xml_state_templates or [])]
    skill_rel = (module_data.get('skill_paths') or [f'skills/{module_id}-analysis/SKILL.md'])[0]
    skill_path = (project_root / skill_rel).resolve() if include_skill else None

    written: List[str] = []
    skipped: List[str] = []
    knowledge_dir.mkdir(parents=True, exist_ok=True)
    (knowledge_dir / 'cases').mkdir(parents=True, exist_ok=True)
    if include_skill:
        assert skill_path is not None
        skill_path.parent.mkdir(parents=True, exist_ok=True)

    _write_json(knowledge_dir / 'module.json', module_data, overwrite, written, skipped)
    _write_json(knowledge_dir / 'subcategories.json', subcategory_items, overwrite, written, skipped)
    _write_json(knowledge_dir / 'log_types.json', [], overwrite, written, skipped)
    _write_jsonl(knowledge_dir / 'evidence_templates.jsonl', evidence_items, overwrite, written, skipped)
    _write_jsonl(knowledge_dir / 'xml_state_templates.jsonl', xml_state_items, overwrite, written, skipped)
    _write_jsonl(knowledge_dir / 'experience_logs.jsonl', [], overwrite, written, skipped)
    _write_jsonl(knowledge_dir / 'cases' / 'case_cards.jsonl', [], overwrite, written, skipped)
    write_evidence_templates_csv(evidence_items, knowledge_dir / 'evidence_templates.csv', overwrite, written, skipped)
    write_evidence_templates_xlsx(evidence_items, knowledge_dir / 'evidence_templates.xlsx', overwrite, written, skipped)
    write_xml_state_templates_csv(xml_state_items, knowledge_dir / 'xml_state_templates.csv', overwrite, written, skipped)
    write_xml_state_templates_xlsx(xml_state_items, knowledge_dir / 'xml_state_templates.xlsx', overwrite, written, skipped)
    _write_text(
        knowledge_dir / 'README.md',
        _knowledge_readme(module_data, subcategory_items),
        overwrite,
        written,
        skipped,
    )
    _write_text(
        knowledge_dir / 'generation_prompt.md',
        build_generation_prompt(module_data, subcategory_items),
        overwrite,
        written,
        skipped,
    )
    if include_skill:
        assert skill_path is not None
        _write_text(skill_path, _skill_draft(module_data, subcategory_items), overwrite, written, skipped)

    loaded = load_project_knowledge_dir(knowledge_dir, project_root=project_root)
    validation_errors = list(loaded.get('errors') or [])
    validation_errors.extend(_detect_encoding_loss(module_data, subcategory_items, evidence_items, knowledge_dir))
    return {
        'project_root': str(project_root),
        'knowledge_dir': str(knowledge_dir),
        'module_id': module_id,
        'module_title': module_data.get('title'),
        'skill_path': str(skill_path) if skill_path else '',
        'written': written,
        'skipped': skipped,
        'validation_errors': validation_errors,
        'loaded_summary': {
            'subcategory_count': len(loaded.get('subcategories') or []),
            'evidence_template_count': len(loaded.get('evidence_templates') or []),
            'xml_state_template_count': len(loaded.get('xml_state_templates') or []),
            'experience_log_count': len(loaded.get('experience_logs') or []),
            'case_count': len(loaded.get('case_cards') or []),
        },
    }


def build_generation_prompt(module: Dict[str, Any], subcategories: List[Dict[str, Any]]) -> str:
    """Return the prompt used by Claude to fill the evidence template draft."""

    module_id = module.get('id') or 'module-id'
    title = module.get('title') or module_id
    sub_lines = []
    for item in subcategories:
        aliases = ', '.join(item.get('aliases') or [])
        alias_text = f' aliases={aliases}' if aliases else ''
        sub_lines.append(f"- {item.get('id')}: {item.get('title')} - {item.get('description')}{alias_text}")
    if not sub_lines:
        sub_lines.append('- unknown: ????????')
    fields = ', '.join(EVIDENCE_TEMPLATE_FIELDS)
    return f"""# Android ??????????

???? `{title}` (`{module_id}`) ?? Android ????????
?????????????????????????????????????? `.claude-web/android-analysis/evidence_templates.csv` ? `evidence_templates.jsonl`?

## ????
{module.get('description') or '(???)'}

## ????
{chr(10).join(sub_lines)}

## ??????
{fields}

## ????
- ?????????????????????????????
- `regex` ????????????????? `TAG + message`?????????????????
- ???????????????????????????????????????????????????
- ???????????trace?meminfo?ANR/tombstone ??????????????? skill ? Deep ????????????????
- `code_location` ???????????????????
- `meaning` ?????????????????????????
- ?? `$package_name` ??????????? `parameters` ???

?????? CSV/JSONL ??????????????? skill ????????????
"""

def _detect_encoding_loss(
    module: Dict[str, Any],
    subcategories: List[Dict[str, Any]],
    evidence_templates: List[Dict[str, Any]],
    knowledge_dir: Path,
) -> List[Dict[str, Any]]:
    """Detect common Windows console mojibake where Chinese text became question marks."""

    checks: List[tuple[str, str]] = []
    for key in ('title', 'description', 'owner_hint'):
        checks.append((f'module.{key}', str(module.get(key) or '')))
    for item in module.get('guide_paths') or []:
        checks.append(('module.guide_paths', str(item)))
    for item in subcategories:
        sid = item.get('id') or '<unknown>'
        checks.append((f'subcategory.{sid}.title', str(item.get('title') or '')))
        checks.append((f'subcategory.{sid}.description', str(item.get('description') or '')))
        for alias in item.get('aliases') or []:
            checks.append((f'subcategory.{sid}.aliases', str(alias)))
    for item in evidence_templates:
        tid = item.get('id') or '<unknown>'
        for key in ('code_location', 'meaning', 'severity'):
            checks.append((f'evidence_template.{tid}.{key}', str(item.get(key) or '')))

    errors: List[Dict[str, Any]] = []
    seen = set()
    for field, text in checks:
        if not _looks_like_encoding_loss(text):
            continue
        key = (field, text[:80])
        if key in seen:
            continue
        seen.add(key)
        errors.append(
            {
                'code': 'possible_encoding_loss',
                'path': str(knowledge_dir),
                'field': field,
                'message': f'{field} contains repeated question marks; check UTF-8 input encoding before using this knowledge pack.',
            }
        )
    return errors


def _looks_like_encoding_loss(text: str) -> bool:
    if not text:
        return False
    # A run of ??? in human text is a strong sign that a Windows console replaced
    # non-ASCII input before Python received it.
    if '???' in text:
        return True
    stripped = text.strip()
    return len(stripped) >= 12 and stripped.count('?') >= 4 and stripped.count('?') / max(len(stripped), 1) > 0.2


def read_evidence_templates_csv(path: Path) -> List[Dict[str, Any]]:
    with Path(path).open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        return [_normalize_evidence_template(row, row.get('module_id') or '') for row in reader]


def write_evidence_templates_csv(
    items: List[Dict[str, Any]],
    path: Path,
    overwrite: bool = True,
    written: List[str] | None = None,
    skipped: List[str] | None = None,
) -> None:
    path = Path(path)
    if _skip_existing(path, overwrite, skipped):
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=EVIDENCE_TEMPLATE_FIELDS)
        writer.writeheader()
        for item in items:
            writer.writerow(_evidence_for_table(item))
    _mark_written(path, written)


def read_evidence_templates_xlsx(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is required for xlsx evidence template import.') from exc
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(v or '').strip() for v in rows[0]]
        out = []
        for row in rows[1:]:
            data = {headers[i]: row[i] if i < len(row) else '' for i in range(len(headers)) if headers[i]}
            if any(str(v or '').strip() for v in data.values()):
                out.append(_normalize_evidence_template(data, str(data.get('module_id') or '')))
        return out
    finally:
        wb.close()


def write_evidence_templates_xlsx(
    items: List[Dict[str, Any]],
    path: Path,
    overwrite: bool = True,
    written: List[str] | None = None,
    skipped: List[str] | None = None,
) -> None:
    if _skip_existing(Path(path), overwrite, skipped):
        return
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is required for xlsx evidence template export.') from exc
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'evidence_templates'
    ws.append(EVIDENCE_TEMPLATE_FIELDS)
    for item in items:
        row = _evidence_for_table(item)
        ws.append([row.get(field, '') for field in EVIDENCE_TEMPLATE_FIELDS])
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max(max(len(str(cell.value or '')) for cell in col) + 2, 12), 48)
    wb.save(path)
    wb.close()
    _mark_written(path, written)


def convert_evidence_templates(knowledge_dir: Path, direction: str, *, overwrite: bool = True) -> Dict[str, Any]:
    """Convert evidence templates between JSONL, CSV, and XLSX for human review."""

    knowledge_dir = Path(knowledge_dir).resolve()
    jsonl_path = knowledge_dir / 'evidence_templates.jsonl'
    csv_path = knowledge_dir / 'evidence_templates.csv'
    xlsx_path = knowledge_dir / 'evidence_templates.xlsx'
    written: List[str] = []
    skipped: List[str] = []
    if direction == 'jsonl_to_csv':
        items = _read_jsonl(jsonl_path)
        write_evidence_templates_csv(items, csv_path, overwrite, written, skipped)
    elif direction == 'csv_to_jsonl':
        items = read_evidence_templates_csv(csv_path)
        _write_jsonl(jsonl_path, items, overwrite, written, skipped)
    elif direction == 'jsonl_to_xlsx':
        items = _read_jsonl(jsonl_path)
        write_evidence_templates_xlsx(items, xlsx_path, overwrite, written, skipped)
    elif direction == 'xlsx_to_jsonl':
        items = read_evidence_templates_xlsx(xlsx_path)
        _write_jsonl(jsonl_path, items, overwrite, written, skipped)
    else:
        raise ValueError(f'Unsupported conversion direction: {direction}')
    loaded = load_project_knowledge_dir(knowledge_dir)
    return {
        'knowledge_dir': str(knowledge_dir),
        'direction': direction,
        'written': written,
        'skipped': skipped,
        'validation_errors': loaded.get('errors') or [],
        'evidence_template_count': len(loaded.get('evidence_templates') or []),
    }


def convert_xml_state_templates(knowledge_dir: Path, direction: str, *, overwrite: bool = True) -> Dict[str, Any]:
    """Convert XML state templates between JSONL, CSV, and XLSX for review."""

    knowledge_dir = Path(knowledge_dir).resolve()
    jsonl_path = knowledge_dir / 'xml_state_templates.jsonl'
    csv_path = knowledge_dir / 'xml_state_templates.csv'
    xlsx_path = knowledge_dir / 'xml_state_templates.xlsx'
    written: List[str] = []
    skipped: List[str] = []
    if direction == 'jsonl_to_csv':
        items = _read_jsonl(jsonl_path)
        write_xml_state_templates_csv(items, csv_path, overwrite, written, skipped)
    elif direction == 'csv_to_jsonl':
        items = read_xml_state_templates_csv(csv_path)
        _write_jsonl(jsonl_path, items, overwrite, written, skipped)
    elif direction == 'jsonl_to_xlsx':
        items = _read_jsonl(jsonl_path)
        write_xml_state_templates_xlsx(items, xlsx_path, overwrite, written, skipped)
    elif direction == 'xlsx_to_jsonl':
        items = read_xml_state_templates_xlsx(xlsx_path)
        _write_jsonl(jsonl_path, items, overwrite, written, skipped)
    else:
        raise ValueError(f'Unsupported conversion direction: {direction}')
    loaded = load_project_knowledge_dir(knowledge_dir)
    return {
        'knowledge_dir': str(knowledge_dir),
        'direction': direction,
        'written': written,
        'skipped': skipped,
        'validation_errors': loaded.get('errors') or [],
        'xml_state_template_count': len(loaded.get('xml_state_templates') or []),
    }


def _normalize_module(module: Dict[str, Any], project_root: Path, include_skill: bool) -> Dict[str, Any]:
    module_id = _clean_id(module.get('id') or project_root.name)
    title = str(module.get('title') or module.get('name') or module_id).strip()
    description = str(module.get('description') or module.get('summary') or f'{title} Android issue analysis module.').strip()
    out = {
        'id': module_id,
        'title': title,
        'description': description,
        'source_roots': _list_str(module.get('source_roots') or ['.']),
        'guide_paths': _list_str(module.get('guide_paths') or _existing_guides(project_root)),
        'default_package_names': _list_str(module.get('default_package_names') or module.get('package_names')),
        'package_resolution': module.get('package_resolution') if isinstance(module.get('package_resolution'), dict) else {'required': False, 'reason': ''},
        'profiles': _list_str(module.get('profiles') or ['functional', 'stability', 'xts', 'memory', 'performance']),
    }
    skill_paths = _list_str(module.get('skill_paths'))
    if include_skill and not skill_paths:
        skill_paths = [f'skills/{module_id}-analysis/SKILL.md']
    out['skill_paths'] = skill_paths
    return out


def _normalize_subcategories(items: List[Dict[str, Any]], module_id: str) -> List[Dict[str, Any]]:
    if not items:
        return [
            {
                'id': 'unknown',
                'module_id': module_id,
                'title': '?????',
                'description': '??????????????????? Claude ????????',
                'aliases': [],
            }
        ]
    out = []
    for item in items:
        sid = _clean_id(item.get('id') or item.get('title') or 'subcategory')
        out.append(
            {
                'id': sid,
                'module_id': str(item.get('module_id') or module_id),
                'title': str(item.get('title') or sid).strip(),
                'description': str(item.get('description') or '').strip(),
                'aliases': _list_str(item.get('aliases')),
            }
        )
    return out

def _normalize_evidence_template(item: Dict[str, Any], module_id: str) -> Dict[str, Any]:
    subcategory_id = item.get('subcategory_id') or item.get('submodule_id') or ''
    return {
        'id': _clean_id(item.get('id')),
        'module_id': str(item.get('module_id') or module_id).strip(),
        'subcategory_id': _clean_id(subcategory_id),
        'profile': str(item.get('profile') or 'functional').strip(),
        'log_type': str(item.get('log_type') or '').strip(),
        'regex': str(item.get('regex') or '').strip(),
        'parameters': _list_str(item.get('parameters')),
        'code_location': str(item.get('code_location') or '').strip(),
        'meaning': str(item.get('meaning') or '').strip(),
        'severity': str(item.get('severity') or 'info').strip(),
        'time_anchor': _bool_value(item.get('time_anchor')),
        'next_steps': _list_str(item.get('next_steps')),
        'enabled': _bool_value(item.get('enabled'), default=True),
    }


def _normalize_xml_state_template(item: Dict[str, Any], module_id: str) -> Dict[str, Any]:
    """Normalize a structured XML/SP state evidence template.

    第一版只做声明式解析：通过路径正则定位 XML/SP 文件，再通过 key/value
    正则提取状态证据。真正执行时由服务端内置解析器处理，不允许运行项目脚本。
    """

    subcategory_id = item.get('subcategory_id') or item.get('submodule_id') or ''
    return {
        'id': _clean_id(item.get('id')),
        'module_id': str(item.get('module_id') or module_id).strip(),
        'subcategory_id': _clean_id(subcategory_id),
        'profile': str(item.get('profile') or 'functional').strip(),
        'source_type': str(item.get('source_type') or 'shared_prefs_xml').strip(),
        'path_patterns': _list_str(item.get('path_patterns')),
        'key_regex': str(item.get('key_regex') or '').strip(),
        'value_regex': str(item.get('value_regex') or '').strip(),
        'value_source': str(item.get('value_source') or 'shared_prefs_value').strip(),
        'code_location': str(item.get('code_location') or '').strip(),
        'meaning': str(item.get('meaning') or '').strip(),
        'severity': str(item.get('severity') or 'info').strip(),
        'time_anchor': _bool_value(item.get('time_anchor')),
        'next_steps': _list_str(item.get('next_steps')),
        'enabled': _bool_value(item.get('enabled'), default=True),
    }


def _evidence_for_table(item: Dict[str, Any]) -> Dict[str, str]:
    normalized = _normalize_evidence_template(item, str(item.get('module_id') or ''))
    out: Dict[str, str] = {}
    for field in EVIDENCE_TEMPLATE_FIELDS:
        value = normalized.get(field)
        if isinstance(value, list):
            out[field] = '; '.join(str(x) for x in value)
        elif isinstance(value, bool):
            out[field] = 'true' if value else 'false'
        else:
            out[field] = str(value or '')
    return out


def _xml_state_for_table(item: Dict[str, Any]) -> Dict[str, str]:
    normalized = _normalize_xml_state_template(item, str(item.get('module_id') or ''))
    out: Dict[str, str] = {}
    for field in XML_STATE_TEMPLATE_FIELDS:
        value = normalized.get(field)
        if isinstance(value, list):
            out[field] = '; '.join(str(x) for x in value)
        elif isinstance(value, bool):
            out[field] = 'true' if value else 'false'
        else:
            out[field] = str(value or '')
    return out


def read_xml_state_templates_csv(path: Path) -> List[Dict[str, Any]]:
    with Path(path).open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        return [_normalize_xml_state_template(row, row.get('module_id') or '') for row in reader]


def write_xml_state_templates_csv(
    items: List[Dict[str, Any]],
    path: Path,
    overwrite: bool = True,
    written: List[str] | None = None,
    skipped: List[str] | None = None,
) -> None:
    path = Path(path)
    if _skip_existing(path, overwrite, skipped):
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=XML_STATE_TEMPLATE_FIELDS)
        writer.writeheader()
        for item in items:
            writer.writerow(_xml_state_for_table(item))
    _mark_written(path, written)


def read_xml_state_templates_xlsx(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is required for xlsx XML state template import.') from exc
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(v or '').strip() for v in rows[0]]
        out = []
        for row in rows[1:]:
            data = {headers[i]: row[i] if i < len(row) else '' for i in range(len(headers)) if headers[i]}
            if any(str(v or '').strip() for v in data.values()):
                out.append(_normalize_xml_state_template(data, str(data.get('module_id') or '')))
        return out
    finally:
        wb.close()


def write_xml_state_templates_xlsx(
    items: List[Dict[str, Any]],
    path: Path,
    overwrite: bool = True,
    written: List[str] | None = None,
    skipped: List[str] | None = None,
) -> None:
    if _skip_existing(Path(path), overwrite, skipped):
        return
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is required for xlsx XML state template export.') from exc
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'xml_state_templates'
    ws.append(XML_STATE_TEMPLATE_FIELDS)
    for item in items:
        row = _xml_state_for_table(item)
        ws.append([row.get(field, '') for field in XML_STATE_TEMPLATE_FIELDS])
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max(max(len(str(cell.value or '')) for cell in col) + 2, 12), 48)
    wb.save(path)
    wb.close()
    _mark_written(path, written)


def _knowledge_readme(module: Dict[str, Any], subcategories: List[Dict[str, Any]]) -> str:
    sub_lines = '\n'.join(f"- `{s['id']}`: {s['title']} - {s.get('description') or ''}" for s in subcategories)
    return f"""# {module.get('title')} Android ?????

???? Claude Web Server ???????????????????????????????????????????????

## ????
1. ??? `module.json` ? `subcategories.json`?
2. ?? `generation_prompt.md` ????????? Claude ??????? `evidence_templates.csv` / `evidence_templates.jsonl`?
3. ??????????? `regex` ????????????????????
4. ? CSV/XLSX ? JSONL ?????????
5. ?????????? `experience_logs.jsonl`????????? `cases/case_cards.jsonl`?

## ????
{sub_lines}
"""

def _skill_draft(module: Dict[str, Any], subcategories: List[Dict[str, Any]]) -> str:
    module_id = module.get('id')
    title = module.get('title')
    sub_lines = '\n'.join(f"- `{s['id']}`: {s['title']} - {s.get('description') or ''}" for s in subcategories)
    return f"""---
name: {module_id}-analysis
description: Analyze {title} Android issues with exact evidence first, then project code and full logs.
---

# {title} ???? Skill

## ????
???? `{title}` (`{module_id}`) ?? Android ?????????????????????????????????????

## ????
{sub_lines}

## ????
1. ??????????????????????????????????
2. ???? `.claude-web/android-analysis/evidence_templates.jsonl` ???????????????
3. ?????????????????`CLAUDE.md` / `AGENTS.md` ??????
4. ????????? `module.json.source_roots`?????????
5. ???????????????????????????

## ??
- ???? grep ???/??/????????????????????????????
- Deep ?????????????????????
- ???????????????????????????
"""

def _write_json(path: Path, data: Any, overwrite: bool, written: List[str], skipped: List[str]) -> None:
    if _skip_existing(path, overwrite, skipped):
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    _mark_written(path, written)


def _write_jsonl(path: Path, items: List[Dict[str, Any]], overwrite: bool, written: List[str], skipped: List[str]) -> None:
    if _skip_existing(path, overwrite, skipped):
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    text = ''.join(json.dumps(item, ensure_ascii=False) + '\n' for item in items)
    path.write_text(text, encoding='utf-8')
    _mark_written(path, written)


def _read_jsonl(path: Path) -> List[Dict[str, Any]]:
    if not Path(path).is_file():
        return []
    out = []
    for line in Path(path).read_text(encoding='utf-8').splitlines():
        text = line.strip()
        if not text or text.startswith('#'):
            continue
        item = json.loads(text)
        if isinstance(item, dict):
            out.append(item)
    return out


def _write_text(path: Path, text: str, overwrite: bool, written: List[str], skipped: List[str]) -> None:
    if _skip_existing(path, overwrite, skipped):
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding='utf-8')
    _mark_written(path, written)


def _skip_existing(path: Path, overwrite: bool, skipped: List[str] | None) -> bool:
    path = Path(path)
    if path.exists() and not overwrite:
        if skipped is not None:
            skipped.append(str(path))
        return True
    return False


def _mark_written(path: Path, written: List[str] | None) -> None:
    if written is not None:
        written.append(str(Path(path)))


def _normalize_relative_path(value: str) -> str:
    value = str(value or DEFAULT_PROJECT_KNOWLEDGE_RELATIVE_PATH).replace('\\', '/').strip().strip('/')
    if not value or value.startswith('../') or '/..' in value or ':' in value:
        return DEFAULT_PROJECT_KNOWLEDGE_RELATIVE_PATH
    return value


def _existing_guides(project_root: Path) -> List[str]:
    return [name for name in ('CLAUDE.md', 'AGENTS.md') if (project_root / name).is_file()]


def _clean_id(value: Any) -> str:
    text = str(value or '').strip()
    if not re.fullmatch(r'[A-Za-z0-9_.:-]+', text):
        text = re.sub(r'[^A-Za-z0-9_.:-]+', '-', text).strip('-')
    return text or 'unknown'


def _list_str(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [p.strip() for p in re.split(r'[;,]\s*|\n+', value) if p.strip()]
    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]
    return [str(value).strip()] if str(value).strip() else []


def _bool_value(value: Any, default: bool = False) -> bool:
    if value is None or value == '':
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}
